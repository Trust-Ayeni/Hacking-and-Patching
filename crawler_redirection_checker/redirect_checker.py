import argparse
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import sys
import logging
import concurrent.futures
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Font
import aiohttp
import asyncio
import asyncio

# Improved logging setup
class ColoredFormatter(logging.Formatter):
    grey = "\x1b[38;21m"
    green = "\x1b[32;21m"
    yellow = "\x1b[33;21m"
    red = "\x1b[31;21m"
    bold_red = "\x1b[31;1m"
    reset = "\x1b[0m"
    format = "%(asctime)s - %(levelname)s - %(message)s"

    FORMATS = {
        logging.DEBUG: grey + format + reset,
        logging.INFO: green + format + reset,
        logging.WARNING: yellow + format + reset,
        logging.ERROR: red + format + reset,
        logging.CRITICAL: bold_red + format + reset
    }

    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno)
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)

handler = logging.StreamHandler()
handler.setFormatter(ColoredFormatter())
logger = logging.getLogger()
logger.addHandler(handler)
logger.setLevel(logging.INFO)

# Session setup with retries and configurable timeout
def create_session():
    session = requests.Session()
    retries = Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    session.mount('http://', HTTPAdapter(max_retries=retries))
    session.mount('https://', HTTPAdapter(max_retries=retries))
    return session

session = create_session()

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:89.0) Gecko/20100101 Firefox/89.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
]

async def fetch(session, url):
    headers = {'User-Agent': random.choice(USER_AGENTS)}
    try:
        async with session.get(url, headers=headers) as response:
            return await response.text(), response.url
    except aiohttp.ClientError as e:
        logger.error(f"Error accessing {url}: {e}")
        return None, None

async def crawl_page_async(url, depth, max_depth, visited, session):
    if url in visited or depth > max_depth:
        return visited

    visited.add(url)
    logger.info(f"Crawling: {url} (depth: {depth})")

    html, final_url = await fetch(session, url)
    if not html:
        return visited

    logger.info(f"Successfully fetched {url} (final URL: {final_url})")

    soup = BeautifulSoup(html, 'lxml')

    tasks = []
    for link in soup.find_all('a'):
        href = link.get('href')
        if href:
            full_url = urljoin(url, href)
            if urlparse(full_url).netloc == urlparse(url).netloc:
                tasks.append(crawl_page_async(full_url, depth + 1, max_depth, visited, session))

    await asyncio.gather(*tasks)
    return visited

def crawl_page(url, depth=0, max_depth=4):
    visited = set()
    async def main():
        async with aiohttp.ClientSession() as session:
            await crawl_page_async(url, depth, max_depth, visited, session)
    asyncio.run(main())
    return visited

def check_redirection(url):
    try:
        start_time = time.time()
        response = session.get(url, allow_redirects=True, timeout=30, headers={'User-Agent': random.choice(USER_AGENTS)})
        end_time = time.time()
        status_code = response.status_code
        final_url = response.url

        if len(response.history) > 0:
            logger.info(f"Redirection detected for {url} (status code: {status_code}, final URL: {final_url}, took {end_time - start_time:.2f} seconds)")
            return final_url, True
        logger.info(f"No redirection for {url} (status code: {status_code}, took {end_time - start_time:.2f} seconds)")
        return url, False
    except requests.RequestException as e:
        logger.error(f"Error accessing {url}: {e}")
        return None, None

def save_results(results, output_file, format):
    logger.info(f"Saving results to {output_file} in {format} format")
    if format == 'txt':
        with open(output_file, 'w') as f:
            for original, final, redirected in results:
                f.write(f"Original URL: {original}\n")
                f.write(f"Final URL: {final}\n")
                f.write(f"Redirected: {redirected}\n")
                f.write("\n")
    elif format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "URL Redirections"

        headers = ["Original URL", "Final URL", "Redirected"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row, (original, final, redirected) in enumerate(results, start=2):
            ws.cell(row=row, column=1, value=original)
            ws.cell(row=row, column=2, value=final)
            ws.cell(row=row, column=3, value=str(redirected))

        wb.save(output_file)

def process_single_url(url, output, format, max_depth):
    logger.info(f"Processing single URL: {url}")
    final_url, redirected = check_redirection(url)
    if final_url:
        logger.info(f"Original URL: {url}")
        logger.info(f"Final URL: {final_url}")
        logger.info(f"Redirected: {redirected}")

        crawled_urls = crawl_page(final_url, max_depth=max_depth)

        if output and format:
            results = [(url, final_url, redirected)]
            for crawled_url in crawled_urls:
                if crawled_url != final_url:
                    results.append((crawled_url, crawled_url, False))
            save_results(results, output, format)

def process_urls_file(file_path, output, format, max_depth):
    logger.info(f"Processing URLs from file: {file_path}")
    results = []
    with open(file_path, 'r') as f:
        urls = f.read().splitlines()

    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        future_to_url = {executor.submit(check_redirection, url): url for url in urls}
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                final_url, redirected = future.result()
                if final_url:
                    logger.info(f"Original URL: {url}")
                    logger.info(f"Final URL: {final_url}")
                    logger.info(f"Redirected: {redirected}")
                    results.append((url, final_url, redirected))

                    crawled_urls = crawl_page(final_url, max_depth=max_depth)
                    for crawled_url in crawled_urls:
                        if crawled_url != final_url:
                            results.append((crawled_url, crawled_url, False))
            except Exception as exc:
                logger.error(f"{url} generated an exception: {exc}")

    if output and format:
        save_results(results, output, format)

def validate_args(args):
    if not args.url and not args.file:
        raise ValueError("Please provide a URL or a file containing URLs.")
    if args.url and args.file:
        raise ValueError("Please provide either a URL or a file, not both.")
    if args.output and args.format not in ['txt', 'xlsx']:
        raise ValueError("Invalid format specified. Choose 'txt' or 'xlsx'.")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description="A script to check for URL redirections and crawl pages.",
        epilog=(
            "Example usage:"
            " python script.py -u http://example.com -d 2,"
            " python script.py -u http://example.com -o results.xlsx -fmt xlsx -d 1,"
            " python script.py -f urls.txt -d 1,"
            " python script.py -f urls.txt -o results.xlsx -fmt xlsx -d 1,"
            " Note:"
            "  For -u, provide a single URL."
            "  For -f, provide a file with one URL per line."
        )
    )
    parser.add_argument('-u', '--url', type=str, help="URL to check for redirections.")
    parser.add_argument('-f', '--file', type=str, help="Path to a file containing URLs to check for redirections.")
    parser.add_argument('-o', '--output', type=str, help="Path to the output file where results will be saved.")
    parser.add_argument('-fmt', '--format', type=str, choices=['txt', 'xlsx'], help="Format for the output file (txt or xlsx).")
    parser.add_argument('-d', '--depth', type=int, default=4, help="Maximum depth for crawling (default: 4)")
    args = parser.parse_args()

    try:
        validate_args(args)
        if args.url:
            process_single_url(args.url, args.output, args.format, args.depth)
        elif args.file:
            process_urls_file(args.file, args.output, args.format, args.depth)
    except ValueError as ve:
        logger.error(f"Argument error: {ve}")
        sys.exit(1)
    except KeyboardInterrupt:
        logger.info("Process interrupted by user. Exiting...")
        sys.exit(1)
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
        sys.exit(1)
